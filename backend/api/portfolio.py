import io
import numpy as np
import pandas as pd
from datetime import date as date_type
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from backend.database import get_db
from backend.db.models import User, Portfolio, Holding, OptimizationResult
from backend.auth.router import get_current_user
from backend.services.market_data import get_multiple_prices, get_price_history
from backend.services.currency import get_ticker_currency, convert
from backend.models.optimizer import optimize_portfolio
from backend.models.bl_optimizer import optimize_black_litterman

router = APIRouter(prefix="/portfolio", tags=["portfolio"])


def _fetch_asset_name(ticker: str) -> str | None:
    try:
        import yfinance as yf
        info = yf.Ticker(ticker).info
        return info.get("longName") or info.get("shortName") or None
    except Exception:
        return None


class HoldingIn(BaseModel):
    ticker: str
    asset_name: Optional[str] = None
    asset_type: str = 'security'
    shares: float
    avg_buy_price: float
    currency: Optional[str] = None
    purchase_date: Optional[str] = None   # ISO date string YYYY-MM-DD
    fees: Optional[float] = 0.0
    notes: Optional[str] = None
    portfolio_id: Optional[int] = None


class HoldingUpdate(BaseModel):
    shares: Optional[float] = None
    avg_buy_price: Optional[float] = None
    purchase_date: Optional[str] = None
    fees: Optional[float] = None
    notes: Optional[str] = None


class PortfolioCreate(BaseModel):
    name: str


class PortfolioUpdate(BaseModel):
    name: Optional[str] = None
    include_in_aggregated: Optional[bool] = None


def _build_holdings_out(holdings, display_currency: str, db=None) -> tuple[list, float]:
    if not holdings:
        return [], 0.0
    # Backfill missing asset names from yfinance (once per holding, then persisted)
    if db is not None:
        for h in holdings:
            if not h.asset_name:
                name = _fetch_asset_name(h.ticker)
                if name:
                    h.asset_name = name
        db.commit()
    prices = get_multiple_prices([h.ticker for h in holdings])
    holdings_out = []
    total_value = 0.0
    for h in holdings:
        price_data = prices.get(h.ticker, {"price": None, "stale": False})
        current_price_native = price_data["price"] or h.avg_buy_price
        price_stale = price_data["stale"]
        native_currency = get_ticker_currency(h.ticker)
        holding_currency = h.currency or display_currency
        value_native = h.shares * current_price_native
        value = convert(value_native, native_currency, display_currency)
        total_value += value
        # avg_buy_price is recorded in the currency of the source file/entry,
        # so convert the live price into that same currency for a fair comparison
        # and to keep avg buy / current price displayed side by side consistently.
        current_price = convert(current_price_native, native_currency, holding_currency)
        pnl_pct = (current_price - h.avg_buy_price) / h.avg_buy_price * 100
        holdings_out.append({
            "id": h.id,
            "portfolio_id": h.portfolio_id,
            "ticker": h.ticker,
            "asset_name": h.asset_name or "",
            "asset_type": h.asset_type,
            "shares": h.shares,
            "avg_buy_price": h.avg_buy_price,
            "currency": holding_currency,
            "current_price": round(current_price, 2),
            "price_stale": price_stale,
            "native_currency": native_currency,
            "value": round(value, 2),
            "pnl_pct": round(pnl_pct, 2),
            "purchase_date": h.purchase_date.isoformat() if h.purchase_date else None,
        })
    return holdings_out, round(total_value, 2)


def _optimize_holdings(holdings: list, risk_score: int, optimize_fn=optimize_portfolio) -> dict:
    """Fetches price history for the given holdings and returns optimized weights."""
    if not holdings:
        raise HTTPException(status_code=400, detail="No holdings to optimize")
    if len(holdings) < 2:
        raise HTTPException(status_code=400, detail="Need at least 2 holdings to optimize")

    unique_tickers = list(dict.fromkeys(h.ticker for h in holdings))
    price_series = {}
    fetch_errors = []
    for ticker in unique_tickers:
        try:
            hist = get_price_history(ticker, period="1y")
            if hist.empty or len(hist) < 10:
                fetch_errors.append(ticker)
                continue
            series = hist["Close"]
            series.index = series.index.normalize().tz_localize(None)
            price_series[ticker] = series
        except Exception as exc:  # noqa: BLE001
            fetch_errors.append(ticker)
            print(f"[optimize] failed to fetch history for {ticker}: {exc}")

    if len(price_series) < 2:
        detail = "Could not fetch enough price history to optimise."
        if fetch_errors:
            detail += f" Failed tickers: {', '.join(fetch_errors)}."
        raise HTTPException(status_code=400, detail=detail)

    prices_df = pd.DataFrame(price_series)
    result = optimize_fn(prices_df, risk_score)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# Must come before /{portfolio_id} routes
@router.get("/metrics")
def portfolio_metrics(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Computes return, volatility, equity share, and diversification for the user's aggregated portfolio.

    Normalisation used by the portfolio radar chart (Radar 2 in AIAdvisor):
      - expected_annual_return_pct : historical annualised weighted return (%)
      - annual_volatility_pct      : annualised portfolio volatility (%)
      - equity_share_pct           : % of total value in equity/ETF holdings
      - n_effective_assets         : 1 / HHI (Herfindahl) — effective number of positions
    """
    portfolios = db.query(Portfolio).filter(
        Portfolio.user_id == current_user.id,
        Portfolio.include_in_aggregated == True,  # noqa: E712
    ).all()
    all_holdings = [h for p in portfolios for h in p.holdings]
    if not all_holdings:
        raise HTTPException(status_code=400, detail="No holdings in aggregated portfolio")

    prices = get_multiple_prices([h.ticker for h in all_holdings])
    values: dict[str, float] = {}
    equity_value = 0.0
    defensive_value = 0.0
    for h in all_holdings:
        price = (prices.get(h.ticker) or {}).get("price") or h.avg_buy_price
        v = h.shares * price
        values[h.ticker] = values.get(h.ticker, 0.0) + v
        if h.asset_type in ("equity", "etf"):
            equity_value += v
        if h.asset_type in ("bond", "cash"):
            defensive_value += v

    total_value = sum(values.values())
    if total_value == 0:
        raise HTTPException(status_code=400, detail="Total portfolio value is zero")

    weights = {t: v / total_value for t, v in values.items()}

    # Fetch 1-year price history for each ticker; skip tickers with too few points
    price_series: dict[str, pd.Series] = {}
    for ticker in weights:
        try:
            hist = get_price_history(ticker, period="1y")
            if not hist.empty and len(hist) >= 20:
                s = hist["Close"]
                s.index = s.index.normalize().tz_localize(None)
                price_series[ticker] = s
        except Exception as exc:  # noqa: BLE001
            print(f"[metrics] failed to fetch history for {ticker}: {exc}")

    if not price_series:
        raise HTTPException(status_code=400, detail="Could not fetch price history for any holding")

    prices_df = pd.DataFrame(price_series).dropna()
    returns_df = prices_df.pct_change().dropna()

    mean_ret_annual = returns_df.mean() * 252
    cov_annual = returns_df.cov() * 252

    # Re-normalise weights to only the tickers that have price history
    w_arr = np.array([weights.get(t, 0.0) for t in prices_df.columns])
    w_sum = w_arr.sum()
    if w_sum > 0:
        w_arr = w_arr / w_sum

    port_return = float(w_arr @ mean_ret_annual.values)
    port_vol = float(np.sqrt(w_arr @ cov_annual.values @ w_arr))

    # Herfindahl diversification index
    hhi = sum(ww ** 2 for ww in weights.values())
    n_eff = 1.0 / hhi if hhi > 0 else 1.0

    return {
        "expected_annual_return_pct": round(port_return * 100, 2),
        "annual_volatility_pct": round(port_vol * 100, 2),
        "equity_share_pct": round(equity_value / total_value * 100, 2),
        "defensive_share_pct": round(defensive_value / total_value * 100, 2),
        "n_effective_assets": round(n_eff, 2),
    }


@router.get("/list")
def list_portfolios(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Retrieves a list of portfolios for the current user, including their holdings and total value."""
    portfolios = db.query(Portfolio).filter(Portfolio.user_id == current_user.id).all()
    display_currency = current_user.display_currency or 'USD'
    result = []
    for p in portfolios:
        _, total = _build_holdings_out(p.holdings, display_currency)
        result.append({
            "id": p.id,
            "name": p.name,
            "holdings_count": len(p.holdings),
            "total_value": total,
            "include_in_aggregated": p.include_in_aggregated,
        })
    return result


@router.get("/")
def get_portfolio_aggregated(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Retrieves the aggregated portfolio for the current user, including holdings and total value in the user's display currency."""
    portfolios = db.query(Portfolio).filter(
        Portfolio.user_id == current_user.id,
        Portfolio.include_in_aggregated == True,  # noqa: E712
    ).all()
    display_currency = current_user.display_currency or 'USD'
    all_holdings = [h for p in portfolios for h in p.holdings]
    holdings_out, total_value = _build_holdings_out(all_holdings, display_currency, db=db)
    return {"holdings": holdings_out, "total_value": total_value, "display_currency": display_currency}


# Must come before /{portfolio_id} routes — "optimize" would otherwise be parsed as portfolio_id
@router.get("/optimize")
def optimize_aggregated(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Optimizes the user's aggregated holdings using Black-Litterman and persists the result."""
    portfolios = db.query(Portfolio).filter(
        Portfolio.user_id == current_user.id,
        Portfolio.include_in_aggregated == True,  # noqa: E712
    ).all()
    all_holdings = [h for p in portfolios for h in p.holdings]
    risk_score = current_user.risk_score or 5
    result = _optimize_holdings(all_holdings, risk_score, optimize_fn=optimize_black_litterman)

    # Persist (upsert) the latest optimization result so exports can read it without recalculating
    existing = db.query(OptimizationResult).filter(
        OptimizationResult.user_id == current_user.id
    ).first()
    if existing:
        existing.weights = result.get("weights", {})
        existing.expected_annual_return_pct = result.get("expected_annual_return_pct")
        existing.annual_volatility_pct = result.get("annual_volatility_pct")
        existing.sharpe_ratio = result.get("sharpe_ratio")
        from datetime import datetime
        existing.created_at = datetime.utcnow()
    else:
        db.add(OptimizationResult(
            user_id=current_user.id,
            weights=result.get("weights", {}),
            expected_annual_return_pct=result.get("expected_annual_return_pct"),
            annual_volatility_pct=result.get("annual_volatility_pct"),
            sharpe_ratio=result.get("sharpe_ratio"),
        ))
    db.commit()

    return result


# ---------------------------------------------------------------------------
# Shared helper: build export rows
# ---------------------------------------------------------------------------

def _build_export_rows(
    current_user: User, db: Session
) -> tuple[list[dict], float, float, OptimizationResult | None]:
    """Return (rows, total_value, total_pnl, opt_result) for export endpoints.

    Each row dict has: ticker, asset_name, asset_type, shares, avg_buy_price,
    currency, current_price, value, pnl_abs, pnl_pct, weight_pct, opt_weight_pct.
    opt_weight_pct is None if the user has never run optimization.
    """
    portfolios = db.query(Portfolio).filter(
        Portfolio.user_id == current_user.id,
        Portfolio.include_in_aggregated == True,  # noqa: E712
    ).all()
    all_holdings = [h for p in portfolios for h in p.holdings]
    if not all_holdings:
        return [], 0.0, 0.0, None

    display_currency = current_user.display_currency or "USD"
    prices = get_multiple_prices([h.ticker for h in all_holdings])

    total_value = 0.0
    raw_rows = []
    for h in all_holdings:
        price_data = prices.get(h.ticker, {"price": None, "stale": False})
        current_price_native = price_data["price"] or h.avg_buy_price
        native_currency = get_ticker_currency(h.ticker)
        holding_currency = h.currency or display_currency
        value = convert(h.shares * current_price_native, native_currency, display_currency)
        current_price = convert(current_price_native, native_currency, holding_currency)
        pnl_abs = (current_price - h.avg_buy_price) * h.shares
        pnl_pct = (current_price - h.avg_buy_price) / h.avg_buy_price * 100 if h.avg_buy_price else 0.0
        total_value += value
        raw_rows.append({
            "ticker": h.ticker,
            "asset_name": h.asset_name or "",
            "asset_type": h.asset_type,
            "shares": h.shares,
            "avg_buy_price": h.avg_buy_price,
            "currency": holding_currency,
            "current_price": round(current_price, 2),
            "value": round(value, 2),
            "pnl_abs": round(pnl_abs, 2),
            "pnl_pct": round(pnl_pct, 2),
        })

    total_pnl = sum(r["pnl_abs"] for r in raw_rows)

    # Optimized weights from last persisted run (None if never optimized)
    opt_result = db.query(OptimizationResult).filter(
        OptimizationResult.user_id == current_user.id
    ).first()
    opt_weights: dict = opt_result.weights if opt_result else {}

    rows = []
    for r in raw_rows:
        r["weight_pct"] = round(r["value"] / total_value * 100, 2) if total_value else 0.0
        opt_w = opt_weights.get(r["ticker"])
        r["opt_weight_pct"] = round(opt_w * 100, 2) if opt_w is not None else None
        rows.append(r)

    return rows, round(total_value, 2), round(total_pnl, 2), opt_result


# ---------------------------------------------------------------------------
# Export — Excel
# ---------------------------------------------------------------------------

@router.get("/export/excel")
def export_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download aggregated portfolio as .xlsx.

    Columns: Ticker, Name, Type, Shares, Avg Buy Price, Currency,
             Current Price, Value, P&L (abs), P&L (%), Weight (%),
             Optimized Weight (%) [blank if never optimized].
    Last row: totals for Value and P&L.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows, total_value, total_pnl, opt_result = _build_export_rows(current_user, db)
    if not rows:
        raise HTTPException(status_code=400, detail="No holdings to export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    headers = [
        "Ticker", "Name", "Type", "Shares",
        "Avg Buy Price", "Currency", "Current Price",
        "Value", "P&L", "P&L %", "Weight %", "Optimized Weight %",
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    display_currency = current_user.display_currency or "USD"
    num_fmt = f'"{display_currency}" #,##0.00'

    for r in rows:
        opt_w = r["opt_weight_pct"] if r["opt_weight_pct"] is not None else ""
        ws.append([
            r["ticker"], r["asset_name"], r["asset_type"], r["shares"],
            r["avg_buy_price"], r["currency"], r["current_price"],
            r["value"], r["pnl_abs"], r["pnl_pct"], r["weight_pct"], opt_w,
        ])
        row_idx = ws.max_row
        for col in (5, 7, 8, 9):   # avg buy, current, value, pnl_abs
            ws.cell(row=row_idx, column=col).number_format = num_fmt
        ws.cell(row=row_idx, column=10).number_format = '0.00"%"'
        ws.cell(row=row_idx, column=11).number_format = '0.00"%"'
        if opt_w != "":
            ws.cell(row=row_idx, column=12).number_format = '0.00"%"'

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 8, total_value).number_format = num_fmt
    ws.cell(total_row, 8).font = Font(bold=True)
    ws.cell(total_row, 9, total_pnl).number_format = num_fmt
    ws.cell(total_row, 9).font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(1, 13):
        ws.cell(total_row, col).fill = total_fill

    # Auto column widths
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date_type.today().isoformat()
    username = (current_user.name or current_user.email).replace(" ", "_")
    filename = f"portfolio_{username}_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Export — PDF
# ---------------------------------------------------------------------------

@router.get("/export/pdf")
def export_pdf(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download aggregated portfolio as a formatted PDF report.

    Sections: header (user + date), holdings table, summary block
    (total value, P&L, and — if a prior optimization exists — return/vol/sharpe),
    disclaimer.
    """
    from fpdf import FPDF

    rows, total_value, total_pnl, opt_result = _build_export_rows(current_user, db)
    if not rows:
        raise HTTPException(status_code=400, detail="No holdings to export")

    display_currency = current_user.display_currency or "USD"
    today = date_type.today().strftime("%d %B %Y")
    username = current_user.name or current_user.email

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    # ── Header ────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Portfolio Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"User: {username}   |   Date: {today}", ln=True, align="C")
    pdf.ln(6)

    # ── Holdings table ────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 9)
    col_widths = [18, 38, 16, 16, 20, 20, 18, 18, 22]
    headers_pdf = ["Ticker", "Name", "Type", "Shares", f"Avg Buy\n({display_currency})",
                   f"Curr Price\n({display_currency})", f"Value\n({display_currency})",
                   "P&L %", "Weight %"]
    # Draw table header
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    for i, (h, w) in enumerate(zip(headers_pdf, col_widths)):
        pdf.multi_cell(w, 8, h, border=1, align="C", fill=True,
                       new_x="RIGHT" if i < len(headers_pdf) - 1 else "LMARGIN",
                       new_y="TOP" if i < len(headers_pdf) - 1 else "NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)

    for i, r in enumerate(rows):
        fill = i % 2 == 0
        pdf.set_fill_color(240, 244, 252) if fill else pdf.set_fill_color(255, 255, 255)
        pnl_color = (0, 150, 80) if r["pnl_pct"] >= 0 else (200, 0, 0)
        values = [
            r["ticker"],
            (r["asset_name"] or "")[:20],
            r["asset_type"],
            f"{r['shares']:.2f}",
            f"{r['avg_buy_price']:.2f}",
            f"{r['current_price']:.2f}",
            f"{r['value']:,.2f}",
            f"{r['pnl_pct']:+.2f}%",
            f"{r['weight_pct']:.2f}%",
        ]
        row_h = 6
        for j, (val, w) in enumerate(zip(values, col_widths)):
            if j == 7:  # P&L % — colour coded
                pdf.set_text_color(*pnl_color)
            else:
                pdf.set_text_color(0, 0, 0)
            is_last = j == len(values) - 1
            pdf.multi_cell(w, row_h, val, border=1, align="C", fill=fill,
                           new_x="RIGHT" if not is_last else "LMARGIN",
                           new_y="TOP" if not is_last else "NEXT")
        pdf.set_text_color(0, 0, 0)

    # ── Summary block ─────────────────────────────────────────────────────────
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Summary", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pnl_sign = "+" if total_pnl >= 0 else ""
    pdf.cell(0, 6, f"Total Portfolio Value:  {display_currency} {total_value:,.2f}", ln=True)
    pdf.cell(0, 6, f"Total P&L:              {display_currency} {pnl_sign}{total_pnl:,.2f}", ln=True)

    if opt_result:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, "Portfolio Metrics  (from last optimization run)", ln=True)
        pdf.set_font("Helvetica", "", 10)
        if opt_result.expected_annual_return_pct is not None:
            pdf.cell(0, 6, f"Expected Annual Return:  {opt_result.expected_annual_return_pct:.2f}%", ln=True)
        if opt_result.annual_volatility_pct is not None:
            pdf.cell(0, 6, f"Annual Volatility:       {opt_result.annual_volatility_pct:.2f}%", ln=True)
        if opt_result.sharpe_ratio is not None:
            pdf.cell(0, 6, f"Sharpe Ratio:            {opt_result.sharpe_ratio:.2f}", ln=True)
        opt_date = opt_result.created_at.strftime("%d %b %Y") if opt_result.created_at else "unknown"
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, f"Optimization run on: {opt_date}", ln=True)
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, "Portfolio metrics not available — run Optimize in the Portfolio page first.", ln=True)

    # ── Disclaimer ────────────────────────────────────────────────────────────
    pdf.ln(8)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(
        0, 5,
        "Disclaimer: This document is for informational purposes only and does not "
        "constitute professional financial advice. Past performance is not indicative "
        "of future results. Always consult a qualified financial advisor before making "
        "investment decisions.",
    )

    buf = io.BytesIO(pdf.output())
    today_iso = date_type.today().isoformat()
    safe_name = username.replace(" ", "_")
    filename = f"portfolio_{safe_name}_{today_iso}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/create", status_code=201)
def create_portfolio(
    data: PortfolioCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Creates a new portfolio for the current user and returns its id and name."""
    portfolio = Portfolio(user_id=current_user.id, name=data.name)
    db.add(portfolio)
    db.commit()
    db.refresh(portfolio)
    return {"id": portfolio.id, "name": portfolio.name}


@router.post("/holdings", status_code=201)
def add_holding(
    data: HoldingIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Adds a new holding to a user's portfolio, creating a default portfolio if none exists."""
    if data.portfolio_id:
        portfolio = db.query(Portfolio).filter(
            Portfolio.id == data.portfolio_id,
            Portfolio.user_id == current_user.id,
        ).first()
        if not portfolio:
            raise HTTPException(status_code=404, detail="Portfolio not found")
    else:
        portfolio = db.query(Portfolio).filter(Portfolio.user_id == current_user.id).first()
        if not portfolio:
            portfolio = Portfolio(user_id=current_user.id, name="My Portfolio")
            db.add(portfolio)
            db.commit()
            db.refresh(portfolio)
    from datetime import date as date_type
    purchase_date = None
    if data.purchase_date:
        try:
            purchase_date = date_type.fromisoformat(data.purchase_date)
        except ValueError:
            pass
    asset_name = data.asset_name or _fetch_asset_name(data.ticker.upper())
    holding = Holding(
        portfolio_id=portfolio.id,
        ticker=data.ticker.upper(),
        asset_name=asset_name,
        asset_type=data.asset_type,
        shares=data.shares,
        avg_buy_price=data.avg_buy_price,
        currency=data.currency or current_user.display_currency,
        purchase_date=purchase_date,
        fees=data.fees or 0.0,
        notes=data.notes,
    )
    db.add(holding)
    db.commit()
    return {"message": "Holding added"}


@router.patch("/holdings/{holding_id}")
def update_holding(
    holding_id: int,
    data: HoldingUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Updates a holding in the database based on the provided data, returning a success message upon completion."""
    user_portfolio_ids = [
        p.id for p in db.query(Portfolio).filter(Portfolio.user_id == current_user.id).all()
    ]
    holding = db.query(Holding).filter(
        Holding.id == holding_id,
        Holding.portfolio_id.in_(user_portfolio_ids),
    ).first()
    if not holding:
        raise HTTPException(status_code=404, detail="Holding not found")
    if data.shares is not None:
        holding.shares = data.shares
    if data.avg_buy_price is not None:
        holding.avg_buy_price = data.avg_buy_price
    if data.purchase_date is not None:
        from datetime import date as date_type
        try:
            holding.purchase_date = date_type.fromisoformat(data.purchase_date)
        except ValueError:
            pass
    if data.fees is not None:
        holding.fees = data.fees
    if data.notes is not None:
        holding.notes = data.notes
    db.commit()
    return {"message": "Holding updated"}


@router.delete("/holdings/{holding_id}")
def delete_holding(
    holding_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Deletes a holding by ID if it belongs to the current user's portfolios, returning a success message."""
    user_portfolio_ids = [
        p.id for p in db.query(Portfolio).filter(Portfolio.user_id == current_user.id).all()
    ]
    holding = db.query(Holding).filter(
        Holding.id == holding_id,
        Holding.portfolio_id.in_(user_portfolio_ids),
    ).first()
    if not holding:
        raise HTTPException(status_code=404, detail="Holding not found")
    db.delete(holding)
    db.commit()
    return {"message": "Holding removed"}


@router.get("/{portfolio_id}")
def get_portfolio_by_id(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retrieves a portfolio by ID for the current user, returning its details and holdings in the user's preferred display currency."""
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    display_currency = current_user.display_currency or 'USD'
    holdings_out, total_value = _build_holdings_out(portfolio.holdings, display_currency, db=db)
    return {
        "id": portfolio.id,
        "name": portfolio.name,
        "holdings": holdings_out,
        "total_value": total_value,
        "display_currency": display_currency,
    }


@router.patch("/{portfolio_id}")
def update_portfolio(
    portfolio_id: int,
    data: PortfolioUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Updates a portfolio with the given ID for the current user, returning the updated portfolio details."""
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    if data.name is not None:
        portfolio.name = data.name
    if data.include_in_aggregated is not None:
        portfolio.include_in_aggregated = data.include_in_aggregated
    db.commit()
    db.refresh(portfolio)
    return {"id": portfolio.id, "name": portfolio.name, "include_in_aggregated": portfolio.include_in_aggregated}


@router.delete("/{portfolio_id}")
def delete_portfolio(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Deletes a portfolio by ID if it belongs to the current user, raising a 404 error if not found."""
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    db.delete(portfolio)
    db.commit()
    return {"message": "Portfolio deleted"}


@router.get("/optimize/{portfolio_id}")
def optimize(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Optimizes a portfolio by fetching price history, calculating optimal weights, and returning the result based on the user's risk score."""
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=400, detail="No holdings to optimize")
    risk_score = current_user.risk_score or 5
    return _optimize_holdings(portfolio.holdings, risk_score)
